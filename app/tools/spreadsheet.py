"""Spreadsheet engine — deterministic .xlsx read/write/aggregate (no LLM).

Uses openpyxl to read cell values, write rows, and compute simple column
aggregates in code — so numeric results are exact, never hallucinated. The model
only chooses the tool and relays the computed result.

Read-only operations are safe (Level 0); writes are Level 1 (draft in workspace).
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional

from app.utils.logger import app_logger


def _cell_value(v: Any):
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


class SpreadsheetTool:
    @classmethod
    def read_sheet(
        cls,
        file_path: str,
        sheet_name: Optional[str] = None,
        limit_rows: int = 200,
    ) -> Dict[str, Any]:
        """Read a sheet into a list of row dicts (deterministic)."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active

            rows = []
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= limit_rows:
                    break
                if headers is None:
                    headers = [(_cell_value(h) if h is not None else f"col_{j}") for j, h in enumerate(row)]
                    continue
                row_dict = {headers[j]: _cell_value(row[j]) for j in range(len(headers)) if j < len(row)}
                rows.append(row_dict)

            wb.close()
            return {"success": True, "sheet": ws.title, "rows": rows, "count": len(rows),
                    "truncated": i >= limit_rows}
        except Exception as e:
            app_logger.warning(f"Spreadsheet read failed: {e}")
            return {"success": False, "error": f"Spreadsheet read failed: {e}"}

    @classmethod
    def write_rows(
        cls,
        file_path: str,
        rows: List[Dict[str, Any]],
        sheet_name: str = "Sheet1",
        overwrite: bool = False,
    ) -> Dict[str, Any]:
        """Write a list of row dicts to a sheet (headers = union of keys)."""
        if not rows or not isinstance(rows, list):
            return {"success": False, "error": "rows must be a non-empty list of dicts."}

        try:
            import openpyxl

            if overwrite:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name
            else:
                wb = openpyxl.load_workbook(file_path) if __import__("os").path.exists(file_path) else openpyxl.Workbook()
                ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

            # Determine header order (preserve first-seen key order across rows).
            headers: List[str] = []
            for r in rows:
                for k in r.keys():
                    if k not in headers:
                        headers.append(k)

            # Write header if the sheet is empty.
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                for j, h in enumerate(headers, 1):
                    ws.cell(1, j, h)

            for r in rows:
                ws.append([r.get(h) for h in headers])

            wb.save(file_path)
            return {"success": True, "path": file_path, "rows_written": len(rows), "headers": headers}
        except Exception as e:
            app_logger.warning(f"Spreadsheet write failed: {e}")
            return {"success": False, "error": f"Spreadsheet write failed: {e}"}

    @classmethod
    def aggregate_column(
        cls,
        file_path: str,
        column: str,
        operation: str = "sum",
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Compute sum/avg/min/max/count on a numeric column (deterministic)."""
        op = (operation or "sum").lower()
        if op not in ("sum", "avg", "min", "max", "count"):
            return {"success": False, "error": f"Unsupported operation '{operation}'. Use sum/avg/min/max/count."}

        data = cls.read_sheet(file_path, sheet_name=sheet_name, limit_rows=100000)
        if not data.get("success"):
            return data

        values = []
        for row in data["rows"]:
            v = row.get(column)
            if isinstance(v, (int, float)):
                values.append(float(v))
            elif isinstance(v, str):
                try:
                    values.append(float(v))
                except ValueError:
                    pass  # skip non-numeric cells

        if op == "count":
            result = len(values)
        elif not values:
            return {"success": False, "error": f"No numeric values found in column '{column}'."}
        elif op == "sum":
            result = sum(values)
        elif op == "avg":
            result = sum(values) / len(values)
        elif op == "min":
            result = min(values)
        else:  # max
            result = max(values)

        return {"success": True, "column": column, "operation": op, "result": round(result, 6), "count": len(values)}
